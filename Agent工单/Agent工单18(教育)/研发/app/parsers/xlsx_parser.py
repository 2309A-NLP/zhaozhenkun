# -*- coding: utf-8 -*-  # 工单18：指定源码编码为 UTF-8。
"""xlsx_parser.py - 工单18智能助教的 XLSX 结构化解析模块。"""  # 工单18：声明当前文件功能。

from __future__ import annotations  # 工单18：启用延迟类型注解。

import io  # 工单18：导入内存流模块。


def parse_xlsx(file_bytes: bytes) -> dict:  # 工单18：解析 XLSX 并保留工作表与行号信息。
    try:  # 工单18：尝试按需导入 Excel 解析依赖。
        from openpyxl import load_workbook  # 工单18：仅在解析 Excel 时导入依赖。
    except ModuleNotFoundError:  # 工单18：处理环境缺少依赖的场景。
        message = "当前运行环境未安装 openpyxl，暂时无法解析 Excel 文件内容。"  # 工单18：定义友好错误提示。
        return {"content_text": message, "chunks": [{"content": message, "summary": message[:60], "location": {"kind": "dependency"}, "modality": "text"}]}  # 工单18：返回提示性结果。
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)  # 工单18：读取工作簿内容。
    lines = []  # 工单18：初始化完整文本列表。
    chunks = []  # 工单18：初始化结构化片段列表。
    for sheet in workbook.worksheets:  # 工单18：遍历全部工作表。
        sheet_lines = []  # 工单18：初始化当前工作表文本列表。
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):  # 工单18：遍历当前工作表全部行。
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]  # 工单18：提取并清洗非空单元格。
            if not values:  # 工单18：跳过空白行。
                continue  # 工单18：继续处理下一行。
            row_line = " | ".join(values)  # 工单18：拼接当前行文本。
            sheet_lines.append(row_line)  # 工单18：写入当前工作表文本列表。
            chunks.append({"content": row_line, "summary": values[0][:60], "location": {"sheet": sheet.title, "row": row_index}, "modality": "table"})  # 工单18：写入表格片段。
        if sheet_lines:  # 工单18：仅当当前工作表有内容时写入整体文本。
            lines.append(f"[Sheet]{sheet.title}\n" + "\n".join(sheet_lines))  # 工单18：写入带工作表标题的文本块。
    return {"content_text": "\n\n".join(lines), "chunks": chunks}  # 工单18：返回 XLSX 解析结果。
